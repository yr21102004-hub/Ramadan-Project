"""
Enhanced Export Utilities for Admin Reports
Supports Excel with formatting, charts, and professional layouts
"""

import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class ExportHelper:
    """Helper class for creating professional export files"""
    
    # Color scheme
    HEADER_COLOR = "D4AF37"  # Gold
    ALT_ROW_COLOR = "F5F5F5"  # Light gray
    BORDER_COLOR = "CCCCCC"  # Gray
    
    @staticmethod
    def create_excel_workbook(title="تقرير"):
        """Create a new Excel workbook with default styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        return wb, ws
    
    @staticmethod
    def style_header_row(ws, row_num, num_cols):
        """Apply professional styling to header row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ExportHelper.HEADER_COLOR, 
                                   end_color=ExportHelper.HEADER_COLOR, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                right=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                top=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                bottom=Side(style='thin', color=ExportHelper.BORDER_COLOR)
            )
    
    @staticmethod
    def style_data_rows(ws, start_row, end_row, num_cols):
        """Apply alternating row colors and borders to data"""
        for row in range(start_row, end_row + 1):
            # Alternate row colors
            fill_color = ExportHelper.ALT_ROW_COLOR if row % 2 == 0 else "FFFFFF"
            
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='Arial', size=11)
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                    right=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                    top=Side(style='thin', color=ExportHelper.BORDER_COLOR),
                    bottom=Side(style='thin', color=ExportHelper.BORDER_COLOR)
                )
    
    @staticmethod
    def auto_adjust_column_width(ws, num_cols):
        """Auto-adjust column widths based on content"""
        for col in range(1, num_cols + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column].width = adjusted_width
    
    @staticmethod
    def add_title_section(ws, title, subtitle=None):
        """Add a professional title section"""
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(name='Arial', size=16, bold=True, color=ExportHelper.HEADER_COLOR)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtitle (date)
        if subtitle:
            ws.merge_cells('A2:G2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(name='Arial', size=10, italic=True)
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return 3 if subtitle else 2  # Return next available row


def export_users_excel(users):
    """Export users to Excel with professional formatting"""
    wb, ws = ExportHelper.create_excel_workbook("المستخدمين")
    
    # Add title
    current_row = ExportHelper.add_title_section(
        ws, 
        "تقرير المستخدمين",
        f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    
    current_row += 1  # Add spacing
    
    # Headers
    headers = ['#', 'اسم المستخدم', 'الاسم الكامل', 'رقم الهاتف', 'البريد الإلكتروني', 
               'موقع المشروع', 'نسبة الإنجاز', 'تاريخ التسجيل']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
    
    ExportHelper.style_header_row(ws, current_row, len(headers))
    
    # Data
    data_start_row = current_row + 1
    for idx, user in enumerate(users, 1):
        row = data_start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.get('username', ''))
        ws.cell(row=row, column=3, value=user.get('full_name', ''))
        ws.cell(row=row, column=4, value=user.get('phone', ''))
        ws.cell(row=row, column=5, value=user.get('email', ''))
        ws.cell(row=row, column=6, value=user.get('project_location', ''))
        ws.cell(row=row, column=7, value=f"{user.get('project_percentage', 0)}%")
        ws.cell(row=row, column=8, value=user.get('created_at', ''))
    
    # Style data rows
    ExportHelper.style_data_rows(ws, data_start_row, data_start_row + len(users) - 1, len(headers))
    
    # Auto-adjust columns
    ExportHelper.auto_adjust_column_width(ws, len(headers))
    
    # Add summary section
    summary_row = data_start_row + len(users) + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "الإحصائيات"
    summary_cell.font = Font(name='Arial', size=12, bold=True, color=ExportHelper.HEADER_COLOR)
    
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="إجمالي المستخدمين:")
    ws.cell(row=summary_row, column=2, value=len(users))
    
    summary_row += 1
    completed = len([u for u in users if u.get('project_percentage', 0) == 100])
    ws.cell(row=summary_row, column=1, value="مشاريع مكتملة:")
    ws.cell(row=summary_row, column=2, value=completed)
    
    summary_row += 1
    ongoing = len([u for u in users if 0 < u.get('project_percentage', 0) < 100])
    ws.cell(row=summary_row, column=1, value="مشاريع جارية:")
    ws.cell(row=summary_row, column=2, value=ongoing)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_payments_excel(payments):
    """Export payments to Excel with professional formatting"""
    wb, ws = ExportHelper.create_excel_workbook("المدفوعات")
    
    # Add title
    current_row = ExportHelper.add_title_section(
        ws, 
        "تقرير المدفوعات",
        f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    
    current_row += 1
    
    # Headers
    headers = ['#', 'اسم المستخدم', 'المبلغ (جنيه)', 'طريقة الدفع', 
               'التفاصيل', 'التاريخ', 'الحالة']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
    
    ExportHelper.style_header_row(ws, current_row, len(headers))
    
    # Data
    data_start_row = current_row + 1
    for idx, payment in enumerate(payments, 1):
        row = data_start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=payment.get('username', ''))
        ws.cell(row=row, column=3, value=float(payment.get('amount', 0)))
        ws.cell(row=row, column=4, value=payment.get('method', ''))
        ws.cell(row=row, column=5, value=payment.get('details', ''))
        ws.cell(row=row, column=6, value=payment.get('timestamp', ''))
        ws.cell(row=row, column=7, value=payment.get('status', 'قيد المراجعة'))
    
    # Style data rows
    ExportHelper.style_data_rows(ws, data_start_row, data_start_row + len(payments) - 1, len(headers))
    
    # Auto-adjust columns
    ExportHelper.auto_adjust_column_width(ws, len(headers))
    
    # Add summary
    summary_row = data_start_row + len(payments) + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "الإحصائيات"
    summary_cell.font = Font(name='Arial', size=12, bold=True, color=ExportHelper.HEADER_COLOR)
    
    summary_row += 1
    total_revenue = sum([float(p.get('amount', 0)) for p in payments])
    ws.cell(row=summary_row, column=1, value="إجمالي الإيرادات:")
    ws.cell(row=summary_row, column=2, value=f"{total_revenue:,.2f} جنيه")
    
    summary_row += 1
    confirmed = len([p for p in payments if p.get('status') == 'Confirmed'])
    ws.cell(row=summary_row, column=1, value="مدفوعات مؤكدة:")
    ws.cell(row=summary_row, column=2, value=confirmed)
    
    summary_row += 1
    pending = len([p for p in payments if 'pending' in p.get('status', '').lower()])
    ws.cell(row=summary_row, column=1, value="قيد المراجعة:")
    ws.cell(row=summary_row, column=2, value=pending)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_chats_excel(chats):
    """Export chat logs to Excel with professional formatting"""
    wb, ws = ExportHelper.create_excel_workbook("سجلات المحادثات")
    
    # Add title
    current_row = ExportHelper.add_title_section(
        ws, 
        "تقرير سجلات المحادثات",
        f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    
    current_row += 1
    
    # Headers
    headers = ['#', 'التوقيت', 'اسم المستخدم', 'الاسم المسجل', 'رسالة العميل', 'رد الذكاء الاصطناعي', 'IP Address']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
    
    ExportHelper.style_header_row(ws, current_row, len(headers))
    
    # Data
    data_start_row = current_row + 1
    for idx, chat in enumerate(chats, 1):
        row = data_start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=chat.get('timestamp', ''))
        ws.cell(row=row, column=3, value=chat.get('username', 'زائر'))
        ws.cell(row=row, column=4, value=chat.get('user_name', ''))
        ws.cell(row=row, column=5, value=chat.get('message', ''))
        ws.cell(row=row, column=6, value=chat.get('response', ''))
        ws.cell(row=row, column=7, value=chat.get('user_ip', ''))
    
    # Style data rows
    ExportHelper.style_data_rows(ws, data_start_row, data_start_row + len(chats) - 1, len(headers))
    
    # Auto-adjust columns (limit width)
    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        # Specific widths for message columns
        if col in [5, 6]: 
            ws.column_dimensions[column].width = 50
        else:
            ws.column_dimensions[column].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_analytics_excel(analytics_data):
    """Export comprehensive analytics to Excel with charts"""
    wb = Workbook()
    
    # ==========================================
    # 1. Overview Sheet
    # ==========================================
    ws1 = wb.active
    ws1.title = "نظرة عامة"
    ws1.sheet_view.rightToLeft = True
    
    current_row = ExportHelper.add_title_section(
        ws1, 
        "تقرير الإحصائيات الشامل",
        f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    current_row += 2
    
    # General stats data
    labels = ['إجمالي المستخدمين', 'إجمالي الرسائل', 'إجمالي المحادثات', 'إجمالي المدفوعات', 
              'طلبات المعاينة', 'الشكاوى', 'أسئلة بدون إجابة', 'مشاريع مكتملة', 'مشاريع جارية']
    values = [
        analytics_data.get('total_users', 0),
        analytics_data.get('total_requests', 0),
        analytics_data.get('total_chats', 0),
        analytics_data.get('total_payments', 0),
        analytics_data.get('total_inspections', 0),
        analytics_data.get('total_complaints', 0),
        analytics_data.get('pending_questions', 0),
        analytics_data.get('completed_projects', 0),
        analytics_data.get('ongoing_projects', 0)
    ]
    
    ws1.cell(row=current_row, column=1, value="المؤشر")
    ws1.cell(row=current_row, column=2, value="القيمة")
    ExportHelper.style_header_row(ws1, current_row, 2)
    
    start_data_row = current_row + 1
    for i, (label, value) in enumerate(zip(labels, values)):
        ws1.cell(row=start_data_row + i, column=1, value=label)
        ws1.cell(row=start_data_row + i, column=2, value=value)
    
    ExportHelper.style_data_rows(ws1, start_data_row, start_data_row + len(labels) - 1, 2)
    ExportHelper.auto_adjust_column_width(ws1, 2)
    
    # Overview Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "نظرة عامة على النظام"
    chart.y_axis.title = "العدد"
    chart.x_axis.title = "المؤشر"
    chart.height = 10
    chart.width = 15
    
    data = Reference(ws1, min_col=2, min_row=start_data_row-1, max_row=start_data_row + len(labels) - 1)
    cats = Reference(ws1, min_col=1, min_row=start_data_row, max_row=start_data_row + len(labels) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws1.add_chart(chart, "D5")

    # ==========================================
    # 2. Services Analysis Sheet
    # ==========================================
    ws_srv = wb.create_sheet("تحليل الخدمات")
    ws_srv.sheet_view.rightToLeft = True
    
    current_row = ExportHelper.add_title_section(ws_srv, "تحليل الخدمات (طلب ومشاهدة)")
    current_row += 1
    
    services_map = {
        'modern-paints': 'دهانات حديثة', 'gypsum-board': 'جبس بورد',
        'integrated-finishing': 'تشطيب متكامل', 'putty-finishing': 'تأسيس ومعجون',
        'wallpaper': 'ورق حائط', 'renovation': 'تجديد وترميم', 'general': 'استفسار عام'
    }
    
    # --- Requested Services ---
    ws_srv.merge_cells(f'A{current_row}:B{current_row}')
    ws_srv[f'A{current_row}'].value = "الخدمات الأكثر طلباً (رسائل)"
    ws_srv[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws_srv.cell(row=current_row, column=1, value="الخدمة")
    ws_srv.cell(row=current_row, column=2, value="عدد الطلبات")
    ExportHelper.style_header_row(ws_srv, current_row, 2)
    
    srv_data = analytics_data.get('services_breakdown', {})
    start_row = current_row + 1
    sorted_srv = sorted(srv_data.items(), key=lambda x: x[1], reverse=True)
    
    for i, (key, count) in enumerate(sorted_srv):
        name = services_map.get(key, key)
        ws_srv.cell(row=start_row + i, column=1, value=name)
        ws_srv.cell(row=start_row + i, column=2, value=count)
        
    if sorted_srv:
        ExportHelper.style_data_rows(ws_srv, start_row, start_row + len(sorted_srv) - 1, 2)
    
    # Chart for Requests
    if sorted_srv:
        pie = PieChart()
        pie.title = "توزيع طلبات الخدمات"
        data = Reference(ws_srv, min_col=2, min_row=start_row-1, max_row=start_row + len(sorted_srv) - 1)
        cats = Reference(ws_srv, min_col=1, min_row=start_row, max_row=start_row + len(sorted_srv) - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.height = 10; pie.width = 10
        ws_srv.add_chart(pie, "D3")
        
    # --- Viewed Services ---
    current_row = start_row + len(sorted_srv) + 5
    ws_srv.merge_cells(f'A{current_row}:B{current_row}')
    ws_srv[f'A{current_row}'].value = "الخدمات الأكثر مشاهدة (زيارات)"
    ws_srv[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws_srv.cell(row=current_row, column=1, value="الخدمة")
    ws_srv.cell(row=current_row, column=2, value="عدد المشاهدات")
    ExportHelper.style_header_row(ws_srv, current_row, 2)
    
    view_data = analytics_data.get('most_viewed_services', {})
    start_row_view = current_row + 1
    sorted_view = sorted(view_data.items(), key=lambda x: x[1], reverse=True)
    
    for i, (key, count) in enumerate(sorted_view):
        ws_srv.cell(row=start_row_view + i, column=1, value=key)
        ws_srv.cell(row=start_row_view + i, column=2, value=count)

    if sorted_view:
        ExportHelper.style_data_rows(ws_srv, start_row_view, start_row_view + len(sorted_view) - 1, 2)
        
    ExportHelper.auto_adjust_column_width(ws_srv, 2)


    # ==========================================
    # 3. Financial Analysis Sheet
    # ==========================================
    ws_fin = wb.create_sheet("التقرير المالي")
    ws_fin.sheet_view.rightToLeft = True
    
    current_row = ExportHelper.add_title_section(ws_fin, "ملخص مالي")
    current_row += 1
    
    ws_fin.cell(row=current_row, column=1, value="المؤشر المالي")
    ws_fin.cell(row=current_row, column=2, value="القيمة")
    ExportHelper.style_header_row(ws_fin, current_row, 2)
    
    fin_labels = ['إجمالي الإيرادات (المؤكدة)', 'عدد عمليات الدفع المؤكدة', 'عمليات قيد المراجعة']
    fin_values = [
        f"{analytics_data.get('total_revenue', 0):,.2f} EGP",
        analytics_data.get('confirmed_payments_count', 0),
        analytics_data.get('pending_payments_count', 0)
    ]
    
    start_row = current_row + 1
    for i, (label, val) in enumerate(zip(fin_labels, fin_values)):
        ws_fin.cell(row=start_row + i, column=1, value=label)
        ws_fin.cell(row=start_row + i, column=2, value=val)
        
    ExportHelper.style_data_rows(ws_fin, start_row, start_row + len(fin_labels) - 1, 2)
    ExportHelper.auto_adjust_column_width(ws_fin, 2)
    
    pie_fin = PieChart()
    pie_fin.title = "حالة المدفوعات"
    # Temp data for financial chart
    ws_fin.cell(row=15, column=1, value="مؤكدة")
    ws_fin.cell(row=15, column=2, value=analytics_data.get('confirmed_payments_count', 0))
    ws_fin.cell(row=16, column=1, value="قيد المراجعة")
    ws_fin.cell(row=16, column=2, value=analytics_data.get('pending_payments_count', 0))
    
    data = Reference(ws_fin, min_col=2, min_row=15, max_row=16)
    cats = Reference(ws_fin, min_col=1, min_row=15, max_row=16)
    pie_fin.add_data(data, titles_from_data=False)
    pie_fin.set_categories(cats)
    ws_fin.add_chart(pie_fin, "D2")

    # ==========================================
    # 4. Advanced Insights Sheet (New)
    # ==========================================
    ws_adv = wb.create_sheet("تحليلات متقدمة")
    ws_adv.sheet_view.rightToLeft = True
    
    current_row = ExportHelper.add_title_section(ws_adv, "تحليلات سلوك المستخدمين المتقدمة")
    current_row += 1
    
    # --- A. Frequent AI Users ---
    ws_adv.merge_cells(f'A{current_row}:B{current_row}')
    ws_adv[f'A{current_row}'].value = "المستخدمين النشطين على AI (يومين فأكثر)"
    ws_adv[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws_adv.cell(row=current_row, column=1, value="المستخدم")
    ws_adv.cell(row=current_row, column=2, value="عدد الأيام النشطة")
    ExportHelper.style_header_row(ws_adv, current_row, 2)
    
    ai_data = analytics_data.get('frequent_ai_users', {})
    start_row = current_row + 1
    # Sort
    sorted_ai = sorted(ai_data.items(), key=lambda x: x[1], reverse=True)
    
    for i, (usr, days) in enumerate(sorted_ai):
        ws_adv.cell(row=start_row + i, column=1, value=usr)
        ws_adv.cell(row=start_row + i, column=2, value=days)
    if sorted_ai:
        ExportHelper.style_data_rows(ws_adv, start_row, start_row + len(sorted_ai) - 1, 2)
    
    current_row = max(start_row + len(sorted_ai), current_row + 1) + 2
    
    # --- B. Daily One-Time Visitors ---
    ws_adv.merge_cells(f'A{current_row}:B{current_row}')
    ws_adv[f'A{current_row}'].value = "زوار اليوم (مرة واحدة)"
    ws_adv[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws_adv.cell(row=current_row, column=1, value="المستخدم")
    ws_adv.cell(row=current_row, column=2, value="عدد الزيارات (1)")
    ExportHelper.style_header_row(ws_adv, current_row, 2)
    
    daily_data = analytics_data.get('daily_one_time_visitors', {})
    start_row = current_row + 1
    for i, (usr, count) in enumerate(daily_data.items()):
        ws_adv.cell(row=start_row + i, column=1, value=usr)
        ws_adv.cell(row=start_row + i, column=2, value=count)
    if daily_data:
        ExportHelper.style_data_rows(ws_adv, start_row, start_row + len(daily_data) - 1, 2)
        
    current_row = max(start_row + len(daily_data), current_row + 1) + 2

    # --- C. All Active Users (Visits) ---
    ws_adv.merge_cells(f'A{current_row}:B{current_row}')
    ws_adv[f'A{current_row}'].value = "جميع المستخدمين النشطين (إجمالي الزيارات)"
    ws_adv[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws_adv.cell(row=current_row, column=1, value="المستخدم")
    ws_adv.cell(row=current_row, column=2, value="إجمالي الزيارات")
    ExportHelper.style_header_row(ws_adv, current_row, 2)
    
    active_data = analytics_data.get('all_active_users', {})
    # Sort
    sorted_active = sorted(active_data.items(), key=lambda x: x[1], reverse=True)
    start_row = current_row + 1
    for i, (usr, count) in enumerate(sorted_active):
        ws_adv.cell(row=start_row + i, column=1, value=usr)
        ws_adv.cell(row=start_row + i, column=2, value=count)
    if sorted_active:
        ExportHelper.style_data_rows(ws_adv, start_row, start_row + len(sorted_active) - 1, 2)

    ExportHelper.auto_adjust_column_width(ws_adv, 2)

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
